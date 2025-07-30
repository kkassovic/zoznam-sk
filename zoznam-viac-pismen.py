import requests
from bs4 import BeautifulSoup
import json
import csv
import os
from openpyxl import Workbook, load_workbook

BASE_URL = "https://www.zoznam.sk"
LETTERS = ["Q", "Y"]  # 👈 Letters to scrape
REGION = "Trenciansky-kraj" # 👈
KRAJ = "6"  # 👈 Region code (fixed for region)

CSV_FILE = "C:\\Users\\karol\\OneDrive\\Desktop\\companies.csv"
EXCEL_FILE = "C:\\Users\\karol\\OneDrive\\Desktop\\companies-trencin-multiple.xlsx"

START_PAGE = 1
END_PAGE = 100 # 👈 tento počet stránok pre jedno písmeno by mal stačiť aj pre "silné" kraje ako napr. bratislavsky. Netreba teda meniť...

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
}

csv_headers = ["url", "name", "adresa", "telefon", "email", "web", "ico"]
file_exists = os.path.isfile(CSV_FILE)
csv_file = open(CSV_FILE, mode='a', newline='', encoding='utf-8')
csv_writer = csv.DictWriter(csv_file, fieldnames=csv_headers)
if not file_exists:
    csv_writer.writeheader()

if os.path.isfile(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(csv_headers)

def scrape_company(link):
    try:
        resp = requests.get(link, headers=headers)
        resp.raise_for_status()
        html_content = resp.content.decode('windows-1250', errors='replace')
        s = BeautifulSoup(html_content, "html.parser")

        script_tag = s.find("script", type="application/ld+json")
        if script_tag:
            json_data = json.loads(script_tag.string)
            name = json_data.get("name", "N/A")
            telephone = json_data.get("telephone", "N/A")
            email = json_data.get("email", "N/A")
            url = json_data.get("url", "N/A")

            address_data = json_data.get("address", {})
            street = address_data.get("streetAddress", "")
            city = address_data.get("addressLocality", "")
            postal = address_data.get("postalCode", "")
            adresa = f"{street}, {postal} {city}".strip(", ")

            ico = json_data.get("@id", "N/A")

            return {
                "url": link,
                "name": name,
                "adresa": adresa,
                "telefon": telephone,
                "email": email,
                "web": url,
                "ico": ico
            }
        else:
            print(f"No JSON-LD data found at {link}")
            return None
    except Exception as e:
        print(f"Error scraping {link}: {e}")
        return None

def scrape_firma_links(url, params=None):
    resp = requests.get(url, headers=headers, params=params)
    resp.raise_for_status()
    html_content = resp.content.decode('windows-1250', errors='replace')
    s = BeautifulSoup(html_content, "html.parser")
    links = [BASE_URL + a["href"] for a in s.find_all("a", href=True) if a["href"].startswith("/firma")]
    return links

# Track summary counts
summary = {}

# Loop through multiple letters
for letter in LETTERS:
    sid = 1173 + (ord(letter) - ord('A'))  # 👈 Dynamically calculate sid
    print(f"\n🔤 Scraping letter: {letter} (sid={sid})")
    letter_count = 0  # Track how many items scraped for this letter

    START_URL = f"{BASE_URL}/katalog/Spravodajstvo-informacie/Abecedny-zoznam-firiem/{letter}/{REGION}.html"
    PAGE_URL = f"{BASE_URL}/katalog/Spravodajstvo-informacie/Abecedny-zoznam-firiem/{letter}/sekcia.fcgi"

    for page in range(START_PAGE, END_PAGE + 1):
        print(f"\n📄 Scraping page {page} for letter {letter}...")
        if page == 1:
            firma_links = scrape_firma_links(START_URL)
        else:
            params = {
                "sid": sid,
                "so": "",
                "page": page,
                "desc": "",
                "shops": "",
                "kraj": KRAJ,
                "okres": "",
                "cast": "",
                "attr": ""
            }
            firma_links = scrape_firma_links(PAGE_URL, params=params)

        if not firma_links:
            print(f"⚠️ No company links found on page {page} for letter {letter}. Stopping this letter.")
            break  # Stop this letter if no more companies found

        print(f"Found {len(firma_links)} company links on page {page}.")

        for idx, link in enumerate(firma_links, 1):
            print(f"  [{idx}] Scraping company: {link}")
            company_data = scrape_company(link)
            if company_data:
                csv_writer.writerow(company_data)
                ws.append([company_data[h] for h in csv_headers])
                letter_count += 1

    summary[letter] = letter_count

# Close CSV and save Excel data
csv_file.close()

# Add summary to a new sheet in Excel
if "Summary" in wb.sheetnames:
    del wb["Summary"]  # Remove old summary if it exists

summary_ws = wb.create_sheet("Summary")
summary_ws.append(["Letter", "Scraped Companies"])  # Header

total_scraped = 0
for letter, count in summary.items():
    summary_ws.append([letter, count])
    total_scraped += count

# Add total row
summary_ws.append(["TOTAL", total_scraped])

# Save Excel
wb.save(EXCEL_FILE)

# Print summary to console
print("\n📊 SUMMARY:")
for letter, count in summary.items():
    print(f"  🔤 {letter}: {count} companies scraped")
print(f"\n✅ TOTAL: {total_scraped} companies scraped across {len(summary)} letters.")
